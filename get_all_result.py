#!/usr/bin/env python
# coding=utf-8
import os
import time
from datetime import datetime

import networkx as nx
import pandas as pd  # 用于生成汇总报告
from openpyxl.styles import Font, PatternFill

from algorithm.classic.louvain import louvain_algorithm
from algorithm.common.util.data_reader.AmazonDataset import AmazonDataset
from algorithm.common.util.data_reader.AmericanFootball import AmericanCollegeFootball
from algorithm.common.util.data_reader.CoraDataset import CoraDataset
from algorithm.common.util.data_reader.dataset_dealer import Dataset
from algorithm.common.util.data_reader.EmailEuCoreDataset import EmailEuCoreDataset
from algorithm.common.util.data_reader.PoliticalBooksDataset import PolbooksDataset
from algorithm.common.util.data_reader.ZKClubDataset import ZKClubDataset
from algorithm.common.util.result_evaluation import CommunityDetectionMetrics


def evaluate_algorithm_on_dataset(dataset_class, algorithm_func, num_runs=10):
    """
    根据 dataset_class 和 algorithm_func 选择对应的数据集和算法，运行多次并取平均值，输出平均评估结果和平均运行时间。
    :param dataset_class: 数据集类 (e.g., ZKClubDataset, EmailEuCoreDataset)
    :param algorithm_func: 算法函数 (e.g., louvain_algorithm, sbm_algorithm, spectral_clustering_algorithm)
    :param num_runs: 运行次数，默认为 10
    :return: dict，包含数据集名称、算法名称、平均运行时间和评估指标
    """
    # 实例化数据集类
    dataset = Dataset(dataset_class())

    # 读取数据集和truthtable
    raw_data, truth_table, number_of_community, dataset_name = dataset.read()
    G = nx.Graph()
    G.add_edges_from(raw_data)

    # 存储多次运行的结果
    runtime_total = 0
    metrics_total = {}

    # 多次运行并累计结果
    for _ in range(num_runs):
        # 记录开始时间
        start_time = time.time()

        # 调用算法并传递社区数量参数
        # if algorithm_func == sbm_algorithm:
        #     communities = algorithm_func(raw_data, num_blocks=number_of_community)
        # elif algorithm_func == spectral_clustering_algorithm:
        #     communities = algorithm_func(raw_data, num_clusters=number_of_community)
        # elif algorithm_func == GCN_train_and_evaluate:
        #     communities = GCN_train_and_evaluate(raw_data, truth_table, device)
        # else:
        #     communities = algorithm_func(raw_data)
        communities = []

        # 记录结束时间
        end_time = time.time()
        runtime = end_time - start_time
        runtime_total += runtime

        # 评估社区划分的表现
        evaluation = CommunityDetectionMetrics(G, communities, truth_table)
        metrics = evaluation.evaluate()

        # 累计每次运行的评估结果
        for metric, value in metrics.items():
            if metric in metrics_total:
                metrics_total[metric] += value
            else:
                metrics_total[metric] = value

    # 计算平均运行时间和平均评估结果
    avg_runtime = runtime_total / num_runs
    avg_metrics = {metric: value / num_runs for metric, value in metrics_total.items()}

    # 收集平均结果并返回
    result = {
        "dataset": dataset_class.__name__,  # 获取类名作为数据集名称
        "algorithm": algorithm_func.__name__,  # 获取函数名作为算法名称
        "runtime": avg_runtime,
    }
    result.update(avg_metrics)  # 将平均评估的结果添加到 result 字典中

    return result


def generate_report(results):
    """
    生成汇总报告并输出为 Excel 文件，并给不同数据集的行设置浅色背景，并加粗显示最高的 NMI、Accuracy、Modularity 和最小的 runtime
    :param results: list of dict, 每次运行结果的集合
    """
    # 使用 pandas 将结果转化为表格形式
    df = pd.DataFrame(results)

    # 获取当前时间
    now = datetime.now()

    # 格式化时间为 2410240118（即 年月日时分）
    formatted_time = now.strftime("%y%m%d%H%M")

    # 保存路径
    file_name = "community_detection_report_" + formatted_time + ".xlsx"
    file_path = os.path.join(r"result\report", file_name)

    # 将结果保存为 Excel 文件
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Results")

        # 获取工作簿和工作表对象
        worksheet = writer.sheets["Results"]

        # 定义不同颜色的填充样式，颜色会在列表中循环使用
        fill_colors = [
            "CCFFFF",  # 浅蓝色
            "CCFFCC",  # 浅绿色
            "FFCCCC",  # 浅红色
            "FFFFCC",  # 浅黄色
            "FFCCFF",  # 浅紫色
        ]
        fill_patterns = [
            PatternFill(start_color=color, end_color=color, fill_type="solid")
            for color in fill_colors
        ]

        # 对数据集相同的行应用不同的背景色，使用颜色循环
        dataset_color_map = {}
        for row in range(
            2, len(df) + 2
        ):  # Excel 的行从 1 开始，而 DataFrame 的行从 0 开始，所以 +2
            current_dataset = worksheet.cell(
                row=row, column=1
            ).value  # 第1列是 dataset 名称

            # 为每个新数据集分配一个颜色
            if current_dataset not in dataset_color_map:
                dataset_color_map[current_dataset] = fill_patterns[
                    len(dataset_color_map) % len(fill_patterns)
                ]

            # 获取当前数据集的颜色
            current_fill = dataset_color_map[current_dataset]

            # 应用颜色到整行
            for col in range(1, len(df.columns) + 1):  # 从第1列开始（Excel的列从1开始）
                worksheet.cell(row=row, column=col).fill = current_fill

        # 获取列名，动态处理 NMI, Accuracy, Modularity 列
        metric_columns = [
            "NMI",
            "Accuracy",
            "Modularity",
            "runtime",
        ]  # 处理 runtime 和其他指标

        # 加粗加亮每个数据集下的 NMI、Accuracy、Modularity 的最大值和 runtime 的最小值
        for dataset_name in df["dataset"].unique():
            dataset_rows = df[df["dataset"] == dataset_name]

            # 找到每个指标的最大/最小值行
            for metric in metric_columns:
                if metric == "runtime":
                    # 最小 runtime 处理
                    min_rows = (
                        dataset_rows[
                            dataset_rows[metric] == dataset_rows[metric].min()
                        ].index
                        + 2
                    )
                else:
                    # 其他指标最大值处理
                    min_rows = (
                        dataset_rows[
                            dataset_rows[metric] == dataset_rows[metric].max()
                        ].index
                        + 2
                    )

                # 获取指标列的索引
                col_index = df.columns.get_loc(metric) + 1

                # 加粗和加亮所有最小/最大值所在的行
                for row in min_rows:
                    worksheet.cell(row=row, column=col_index).font = Font(bold=True)
                    worksheet.cell(row=row, column=col_index).fill = PatternFill(
                        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                    )  # 黄色高亮

    print(f"\n汇总报告已保存为 {file_path}")


# 示例调用
if __name__ == "__main__":
    # 列出所有的数据集和算法组合，直接使用类和函数而不是字符串
    dataset_classes = [
        ZKClubDataset,
        PolbooksDataset,
        AmericanCollegeFootball,
        EmailEuCoreDataset,
        CoraDataset,
        AmazonDataset,
    ]  # 可以继续添加新数据集
    # dataset_classes = [ZKClubDataset, PolbooksDataset, AmericanCollegeFootball]
    algorithm_functions = [louvain_algorithm]  # 可以继续添加新算法

    # 存储所有运行的结果
    results = []

    # 循环遍历所有的数据集和算法组合
    for dataset_class in dataset_classes:
        for algorithm_func in algorithm_functions:
            print(
                f"\n==== 运行 {algorithm_func.__name__} 算法在 {dataset_class.__name__} 数据集上 ====\n"
            )
            try:
                result = evaluate_algorithm_on_dataset(
                    dataset_class, algorithm_func, num_runs=1
                )
                results.append(result)  # 收集每次的平均结果
            except Exception as e:
                print(
                    f"运行 {algorithm_func.__name__} 在 {dataset_class.__name__} 时出错: {e}"
                )

    # 生成汇总报告
    generate_report(results)
